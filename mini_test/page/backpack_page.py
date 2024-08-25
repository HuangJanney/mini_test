
from page.main_page import MainPage
from utils.log_util import logger
from airtest.core.api import *
from airtest.aircv import *
from PIL import Image
from PIL import ImageChops

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image


class BackpackPage(MainPage):

    #获取背包第一格物品icon
    def get_first_icon(self,filepath,*subdir):
        backpack_snapshot = self.get_snapshot_path('backpack_snapshot.png')
        self.snapshot_img('backpack_snapshot.png')
        self.sleep(2)
        first_icon = self.cut_img(backpack_snapshot,1467,206,1621,362,filepath,*subdir)
        return first_icon

    #背包手持第一格道具
    def handheld_prop(self):
        self.open_backpack()
        self.sleep(1.0)
        self.touch_coordinate(0.64, 0.268)
        print("手持道具")
        self.sleep()
        self.touch_back()
        if self.exists_recursively_default('main_ui','x.png'):
            logger.info("手持成功")

    #使用第一格背包的道具
    def use_first_prop(self):
        if self.exists_recursively_default('main_ui','x.png'):
            self.touch_recursively_default('main_ui','x.png')
            self.handheld_prop()
            self.touch_coordinate(0.89,0.77)
            wait_time = 0
            while wait_time < 4:
                try:
                    flag = bool(self.exists_recursively_default('main_ui','x.png'))
                    logger.info(flag)
                    if self.exists_recursively_default('main_ui','tips.png'):
                        self.error_report()
                        break
                    elif flag == True:
                        self.sleep(1)
                        logger.info("道具使用中，请等待...")
                    elif flag == False:
                        logger.info('已使用完')
                        break
                except:
                    logger.info('使用异常')
                wait_time += 1
        else:
            self.handheld_prop()
            self.touch_coordinate(0.89,0.77)
            wait_time = 0
            while wait_time < 4:
                try:
                    flag = bool(self.exists_recursively_default('main_ui','x.png'))
                    logger.info(flag)
                    if self.exists_recursively_default('main_ui','tips.png'):
                        self.error_report()
                        break
                    elif flag == True:
                        self.sleep(1)
                        logger.info("道具使用中，请等待...")
                    elif flag == False:
                        logger.info('已使用完')
                        break
                except:
                    logger.info('使用异常')
                wait_time += 1  

    #清空背包
    def clear_backpack(self):
        self.gm_cmd('clean_backpack')
    
    #添加物品GM
    def add_backpack(self,item_id,num):
        self.gm_cmd('add_backpack %s %s'%(item_id,num))


    #图片对比（后续挪到image_util）
    def is_similar(self,img1,img2):
        img1_1 = aircv.imread(img1)
        img2_2 = aircv.imread(img2)
        similarity = find_template(img1_1,img2_2,-1)['confidence']
        print("相似度：",similarity)
        if similarity > 0.7:
            return True
        else:
            return False


    def write_to_excel(self, value, bimg, aimg,result,name):
        filename = f"测试结果_{name}.xlsx"
        if not os.path.exists(filename):
            # 创建一个新的Excel工作簿
            workbook = Workbook()
            sheet = workbook.active

            # 写入列标题
            sheet['A1'] = 'value'
            sheet['B1'] = 'bimg'
            sheet['C1'] = 'aimg'
            sheet['D1'] = 'result'
        else:
            # 打开已存在的Excel文件
            workbook = load_workbook(filename)
            sheet = workbook.active

        # 写入测试结果
        row = sheet.max_row + 1
        sheet.cell(row=row, column=1, value=value)
        sheet.cell(row=row, column=2, value=bimg)
        sheet.cell(row=row, column=3, value=aimg)
        sheet.cell(row=row, column=4, value=result)

        # 插入bimg图片
        bimg_cell = sheet.cell(row=row, column=2)
        bimg_obj = Image(bimg)
        bimg_obj.width = 100  # 调整图片宽高
        bimg_obj.height = 100  # 调整图片高度
        sheet.add_image(bimg_obj, bimg_cell.coordinate)

        # 插入aimg图片
        aimg_cell = sheet.cell(row=row, column=3)
        aimg_obj = Image(aimg)
        aimg_obj.width = 100  # 调整图片宽度
        aimg_obj.height = 100  # 调整图片高度
        sheet.add_image(aimg_obj, aimg_cell.coordinate)

        # 保存Excel文件
        workbook.save(filename)


    #开启制造台
    def open_craft_station(self):
        if self.exists_recursively_default('main_ui','manufacture_icon.png'):
            self.touch_recursively_default('main_ui','manufacture_icon.png')
        else:
            self.operate_reinforce(True)
            if self.exists_recursively_default('main_ui','manufacture_icon.png'):
                self.touch_recursively_default('main_ui','manufacture_icon.png')
            else:
                raise Exception('Not Find Craft Station')
            
    #制造列表向上滑动并点击
    def find_and_touch_Craftable(self,*recursive_path):
        scroll_times = 5
        while scroll_times > 0:
            pos = self.exists_recursively(0.8, *recursive_path)
            if pos:
                touch(pos)
                return
            self.swipe([0.82, 0.42], [0.82, 0.25])
            scroll_times -= 1
            sleep(2)

    #右侧页签滑动
    def scroll_right_tap(self,toward='up'):
        if toward == 'up':
            self.swipe([0.96,0.62],[0.96,0.28])
        elif toward == 'down':
            self.swipe([0.96,0.28],[0.96,0.62])

    #切换页签
    def switch_craft_tap(self,tap):
        """    
        :param tap= 0:Favorited 1:Craftable 2:Tool 3:Relic 4:Furniture 5:Small_Object 6:Wall_Ornament 7:Wallpaper 8:Others
        """
        tap_name_list = ['Favorited','Craftable','Tool','Relic','Furniture','Small_Object','Wall_Ornament','Wallpaper','Others']
        if type(tap) == int:
            if tap < 7 :
                self.scroll_right_tap('up')
                self.touch_recursively_default('fabricate','%s.png'%tap_name_list[tap])
            else:
                self.scroll_right_tap('down')
                self.touch_recursively_default('fabricate','%s.png'%tap_name_list[tap])
        

    #制造物品
    def craft(self,go_on=True):
        if self.exists_recursively_default('fabricate','Craft.png'):
            self.touch_recursively_default('fabricate','Craft.png')
            self.wait_recursively_default('fabricate','Continue.png')
            if go_on:
                self.touch_recursively_default('fabricate','Continue.png')
            else:
                self.touch_recursively_default('fabricate','Quit.png')
        elif self.exists_recursively_default('fabricate','insufficient_materials.png'):
            raise Exception('Insufficient materials')
        

    #制造物品打包
    def craft_finish(self,*recursive_path):
        self.open_craft_station()
        self.switch_craft_tap(1)
        self.find_and_touch_Craftable(*recursive_path)
        self.craft(False)

